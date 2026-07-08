import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import anthropic
import json
import re
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")  
GMAIL_USER = os.environ.get("GMAIL_USER", "marcotullio.valiante@gmail.com")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
EMAIL_DESTINATARIO = os.environ.get("EMAIL_DESTINATARIO", "")

def carica_destinatari():
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            config = json.load(f)
            return config.get("destinatari", [EMAIL_DESTINATARIO])
    except:
        return [EMAIL_DESTINATARIO]

FILE_BANDI_VISTI = "bandi_visti.json"

def analizza_bando_con_claude(titolo, fonte, scadenza):
    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        messaggio = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            messages=[{
                "role": "user",
                "content": f"""Analizza questo bando pubblico italiano. Rispondi SOLO con un oggetto JSON valido, niente altro.

Titolo: {titolo}
Fonte: {fonte}
Scadenza: {scadenza}

Struttura JSON richiesta:
{{
  "tipo_beneficiario": "...",
  "settore_ateco": "...",
  "settore_ets": "...",
  "fascia_demografica_target": "...",
  "qualita_bando": "...",
  "motivazione": "..."
}}

Istruzioni per ogni campo:

"tipo_beneficiario": Chi può accedere al bando. Scegli UNO o PIÙ tra: Aziende, Enti Locali, Terzo Settore, Privati. Se più categorie, separale con " | " (es. "Aziende | Terzo Settore"). Se non determinabile: "N/D".

"settore_ateco": Compilare SOLO se tipo_beneficiario include "Aziende". Scegli la sezione ATECO 2025 più pertinente tra: A-Agricoltura, B-Estrazione, C-Manifatturiero, D-Energia, E-Ambiente, F-Costruzioni, G-Commercio, H-Trasporti, I-Turismo/Ristorazione, J-Comunicazione, K-Finanza, L-Immobiliare, M-Professionale, N-Amministrativo, O-PA, P-Istruzione, Q-Sanità, R-Arte/Sport, S-Altri servizi, T-Famiglie, U-Organizzazioni internazionali. Se non applicabile o non determinabile: "N/A".

"settore_ets": Compilare SOLO se tipo_beneficiario include "Terzo Settore". Scegli tra le attività art. 5 CTS: Assistenza sociale, Assistenza sanitaria, Educazione/istruzione, Tutela ambiente, Protezione civile, Cultura/arte, Tutela diritti, Ricerca scientifica, Sport dilettantistico, Cooperazione internazionale, Inclusione sociale, Agricoltura sociale, Attività commerciali ETS. Se non applicabile: "N/A".

"fascia_demografica_target": Compilare SOLO se tipo_beneficiario include "Enti Locali" E il bando ha soglie demografiche. Scegli tra: Fino a 5.000 abitanti, 5.001-15.000, 15.001-50.000, Oltre 50.000, Tutti i Comuni. Se non applicabile: "N/A".

"qualita_bando": Valuta la qualità intrinseca e accessibilità del bando indipendentemente dal settore. Alta = dotazione rilevante, requisiti chiari, procedura semplice. Media = buon bando ma con complessità o requisiti restrittivi. Bassa = dotazione limitata, iter complesso, o requisiti molto selettivi. Scegli: Alta, Media, Bassa.

"motivazione": 1-2 frasi concrete sul perché questo bando è rilevante per il territorio del Cilento/Campania, citando il tipo di spesa finanziabile o il beneficio principale. Evita frasi generiche."""
            }]
        )
        testo = messaggio.content[0].text.strip()
        match = re.search(r'\{.*\}', testo, re.DOTALL)
        if match:
            return json.loads(match.group())
        return {
            "tipo_beneficiario": "N/D", "settore_ateco": "N/A",
            "settore_ets": "N/A", "fascia_demografica_target": "N/A",
            "qualita_bando": "N/D", "motivazione": testo[:150]
        }
    except Exception as e:
        return {
            "tipo_beneficiario": "N/D", "settore_ateco": "N/A",
            "settore_ets": "N/A", "fascia_demografica_target": "N/A",
            "qualita_bando": "N/D", "motivazione": str(e)
        }

BASE_URL_CAMPANIA = "https://agricoltura.regione.campania.it/"

def parse_data(testo):
    if not testo or testo == "Non specificata":
        return None
    mesi = {
        "gennaio": "01", "febbraio": "02", "marzo": "03",
        "aprile": "04", "maggio": "05", "giugno": "06",
        "luglio": "07", "agosto": "08", "settembre": "09",
        "ottobre": "10", "novembre": "11", "dicembre": "12"
    }
    testo_pulito = testo.lower().strip().split(" -")[0].split("–")[0].strip()
    testo_pulito = testo_pulito.replace("prorogato al", "").strip()
    for mese_it, mese_num in mesi.items():
        if mese_it in testo_pulito:
            testo_pulito = testo_pulito.replace(mese_it, mese_num)
    formati = ["%d/%m/%Y", "%d %m %Y", "%d %m%Y", "%Y-%m-%d"]
    for fmt in formati:
        try:
            return datetime.strptime(testo_pulito.strip(), fmt)
        except:
            continue
    return None

def calcola_stato(scadenza_testo):
    data = parse_data(scadenza_testo)
    if data is None:
        return "Non specificata"
    return "✅ Aperto" if data >= datetime.today() else "❌ Scaduto"

def scrapa_regione_campania():
    try:
        url = BASE_URL_CAMPANIA + "bandi.html"
        risposta = requests.get(url, timeout=10)
        soup = BeautifulSoup(risposta.text, "html.parser")
        bandi = []
        tabelle = soup.find_all("table", class_="table")
        for tabella in tabelle:
            righe = tabella.find_all("tr")
            for riga in righe[1:]:
                celle = riga.find_all("td")
                if len(celle) >= 2:
                    titolo = celle[0].text.strip()
                    scadenza = celle[1].text.strip()
                    link_tag = celle[2].find("a") if len(celle) > 2 else None
                    link_rel = link_tag.get("href", "") if link_tag else ""
                    link = BASE_URL_CAMPANIA + link_rel if link_rel else ""
                    if titolo and titolo != "nessun bando aperto":
                        bandi.append({
                            "Titolo": titolo,
                            "Scadenza": scadenza,
                            "Data pubblicazione": "Non specificata",
                            "Link": link,
                            "Fonte": "Regione Campania - Agricoltura"
                        })
        return bandi
    except Exception as e:
        print(f"  ⚠️ Errore Regione Campania: {e}")
        return []

def scrapa_invitalia():
    try:
        url = "https://www.invitalia.it/per-le-imprese/incentivi-e-strumenti"
        headers = {"User-Agent": "Mozilla/5.0"}
        risposta = requests.get(url, timeout=10, headers=headers)
        soup = BeautifulSoup(risposta.text, "html.parser")
        bandi = []

        for card in soup.find_all("div", class_="card-body"):
            # Titolo
            titolo_tag = card.find("h3")
            if not titolo_tag:
                continue
            titolo = titolo_tag.text.strip()

            # Link diretto — cerca prima read-more, poi qualsiasi link con /incentivi-e-strumenti/
            link_tag = card.find("a", class_=lambda x: x and "read-more" in x)
            if link_tag:
                href = link_tag.get("href", "")
                link = f"https://www.invitalia.it{href}" if href.startswith("/") else href
            else:
                link_tag = card.find("a", href=lambda x: x and "/incentivi-e-strumenti/" in x)
                if link_tag:
                    href = link_tag.get("href", "")
                    link = f"https://www.invitalia.it{href}" if href.startswith("/") else href
                else:
                    link = url

            # Date
            scadenza = "Non specificata"
            date_div = card.find("p", class_=lambda x: x and "dateContainer" in x)
            if date_div:
                testo_date = date_div.text.strip()
                if "chiusura" in testo_date.lower():
                    scadenza = testo_date.split("chiusura:")[-1].strip() if "chiusura:" in testo_date.lower() else testo_date

            if titolo:
                bandi.append({
                    "Titolo": titolo,
                    "Scadenza": scadenza,
                    "Data pubblicazione": "Non specificata",
                    "Link": link,
                    "Fonte": "Invitalia"
                })
        return bandi
    except Exception as e:
        print(f"  ⚠️ Errore Invitalia: {e}")
        return []

def scrapa_gal_cilento():
    try:
        url = "https://www.galcilento.it/bandi/"
        risposta = requests.get(url, timeout=10)
        soup = BeautifulSoup(risposta.text, "html.parser")
        bandi = []
        for bando_div in soup.find_all("div", class_="bando-single"):
            titolo_tag = bando_div.find("h2")
            titolo = titolo_tag.text.strip() if titolo_tag else ""
            link_tag = bando_div.find("a")
            link = link_tag.get("href", "") if link_tag else ""
            chiusura = ""
            for te in bando_div.find_all("div", class_="time-event"):
                testo = te.text.strip()
                if "Chiusura" in testo:
                    chiusura = testo.replace("Chiusura:", "").strip()
            if titolo:
                bandi.append({
                    "Titolo": titolo,
                    "Scadenza": chiusura if chiusura else "Non specificata",
                    "Data pubblicazione": "Non specificata",
                    "Link": link,
                    "Fonte": "GAL Cilento"
                })
        return bandi
    except Exception as e:
        print(f"  ⚠️ Errore GAL Cilento: {e}")
        return []

def scrapa_bandi_ue():
    try:
        print("  Scaricando database bandi UE (attendere...)")
        url = "https://ec.europa.eu/info/funding-tenders/opportunities/data/referenceData/grantsTenders.json"
        risposta = requests.get(url, timeout=60)
        data = risposta.json()
        tutti = data.get("fundingData", {}).get("GrantTenderObj", [])
        bandi = []
        for b in tutti:
            status = b.get("status", {}).get("abbreviation", "")
            if status != "Open":
                continue
            titolo = b.get("title", "N/D")
            deadline_list = b.get("deadlineDatesLong", [])
            if deadline_list:
                scadenza = datetime.fromtimestamp(deadline_list[-1] / 1000).strftime("%d/%m/%Y")
            else:
                scadenza = "Non specificata"
            identificatore = b.get("identifier", "")
            link = f"https://ec.europa.eu/info/funding-tenders/opportunities/portal/screen/opportunities/topic-details/{identificatore}"
            programma = b.get("frameworkProgramme", {}).get("abbreviation", "UE")
            pub_date_ts = b.get("publicationDateLong", None)
            if pub_date_ts:
                data_pub = datetime.fromtimestamp(pub_date_ts / 1000).strftime("%d/%m/%Y")
            else:
                data_pub = "Non specificata"
            bandi.append({
                "Titolo": titolo,
                "Scadenza": scadenza,
                "Data pubblicazione": data_pub,
                "Link": link,
                "Fonte": f"UE - {programma}"
            })
        return bandi
    except Exception as e:
        print(f"  ⚠️ Errore bandi UE: {e}")
        return []

def scrapa_incentivi_gov():
    url = "https://www.incentivi.gov.it/solr/coredrupal/select?q.op=OR&wt=json&rows=8000&fl=nid%3Azs_nid%2Cpage_title%3Azs_title%2Copen_date%3Azs_field_open_date%2Cclose_date%3Azs_field_close_date%2Cregions%3Azm_field_regions%2Csubject_type%3Azm_field_subject_type%2Csupport_form%3Azm_field_support_form%2C&q=index_id%3Aincentivi&sort=ds_last_update+desc"
    headers = {"User-Agent": "Mozilla/5.0 (compatible; LumenScout/1.0)"}
    
    for tentativo in range(1, 4):
        try:
            print(f"  Scaricando incentivi da incentivi.gov.it (tentativo {tentativo}/3)...")
            risposta = requests.get(url, timeout=60, headers=headers)
            data = risposta.json()
            docs = data["response"]["docs"]
            oggi = datetime.today()
            bandi = []
            for doc in docs:
                titolo = doc.get("page_title", "N/D")
                close_date = doc.get("close_date", "")
                open_date = doc.get("open_date", "")
                nid = doc.get("nid", "")
                link = f"https://www.incentivi.gov.it/it/catalogo/{nid}"
                if close_date:
                    scadenza_dt = datetime.strptime(close_date[:10], "%Y-%m-%d")
                    if scadenza_dt < oggi:
                        continue
                    scadenza = scadenza_dt.strftime("%d/%m/%Y")
                else:
                    scadenza = "Non specificata"
                if open_date:
                    data_pub = datetime.strptime(open_date[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                else:
                    data_pub = "Non specificata"
                bandi.append({
                    "Titolo": titolo,
                    "Scadenza": scadenza,
                    "Data pubblicazione": data_pub,
                    "Link": link,
                    "Fonte": "incentivi.gov.it"
                })
            print(f"  ✅ incentivi.gov.it: {len(bandi)} bandi trovati")
            return bandi
        except Exception as e:
            print(f"  ⚠️ Tentativo {tentativo} fallito: {e}")
            if tentativo < 3:
                import time
                time.sleep(10 * tentativo)
    print("  ⚠️ incentivi.gov.it non raggiungibile dopo 3 tentativi — fonte saltata")
    return []

def carica_bandi_visti():
    try:
        with open(FILE_BANDI_VISTI, "r", encoding="utf-8") as f:
            return set(json.load(f))
    except:
        return set()

def salva_bandi_visti(titoli):
    with open(FILE_BANDI_VISTI, "w", encoding="utf-8") as f:
        json.dump(list(titoli), f, ensure_ascii=False)

def invia_email_bandi(bandi_nuovi, totale_aperti):
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"🔔 Lumen Scout — {len(bandi_nuovi)} nuovi bandi — {datetime.now().strftime('%d/%m/%Y')}"
        msg["From"] = GMAIL_USER
        destinatari = carica_destinatari()
        msg["To"] = ", ".join(destinatari)

        righe = ""
        for b in bandi_nuovi[:20]:
            qualita = b.get("Qualità Bando", "")
            colore = "#166534" if qualita == "Alta" else "#92400e" if qualita == "Media" else "#6B7280"
            righe += f"""
            <tr>
                <td style="padding:8px;border-bottom:1px solid #e5e7eb;">{b['Titolo'][:80]}</td>
                <td style="padding:8px;border-bottom:1px solid #e5e7eb;">{b['Scadenza']}</td>
                <td style="padding:8px;border-bottom:1px solid #e5e7eb;">{b.get('Tipo Beneficiario', 'N/D')}</td>
                <td style="padding:8px;border-bottom:1px solid #e5e7eb;color:{colore};font-weight:bold">{qualita}</td>
                <td style="padding:8px;border-bottom:1px solid #e5e7eb;">{b['Fonte']}</td>
            </tr>"""

        html = f"""
        <html><body style="font-family:Arial,sans-serif;max-width:800px;margin:0 auto">
            <div style="background:#0a2e22;color:white;padding:20px;border-radius:8px 8px 0 0">
                <h2 style="margin:0">🔔 Lumen Scout</h2>
                <p style="margin:3px 0 0 0;color:#C9A84C;font-size:13px">Lumen Opportunities | Lumen Advisors</p>
                <p style="margin:5px 0 0 0">{datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
            </div>
            <div style="background:#f9fafb;padding:20px;border-radius:0 0 8px 8px">
                <p>Sono stati trovati <strong>{len(bandi_nuovi)} nuovi bandi</strong> oggi.<br>
                Totale bandi aperti nel sistema: <strong>{totale_aperti}</strong></p>
                <table style="width:100%;border-collapse:collapse;background:white;border-radius:8px">
                    <tr style="background:#0F6E56;color:white">
                        <th style="padding:10px;text-align:left">Titolo</th>
                        <th style="padding:10px;text-align:left">Scadenza</th>
                        <th style="padding:10px;text-align:left">Beneficiari</th>
                        <th style="padding:10px;text-align:left">Qualità</th>
                        <th style="padding:10px;text-align:left">Fonte</th>
                    </tr>
                    {righe}
                </table>
                <p style="color:#6B7280;font-size:12px;margin-top:20px">
                    Lumen Scout — Lumen Opportunities | Lumen Advisors — aggiornamento automatico giornaliero
                </p>
            </div>
        </body></html>"""

        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD.replace(" ", ""))
            server.sendmail(GMAIL_USER, destinatari, msg.as_string())
        print(f"✅ Email inviata a {', '.join(destinatari)}")
    except Exception as e:
        print(f"⚠️ Errore invio email: {e}")

def formatta_excel(filename, n_aperti, n_ns, n_scaduti):
    wb = load_workbook(filename)
    ws = wb.active
    ws.title = "Bandi"
    header_fill = PatternFill("solid", fgColor="0F6E56")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    aperto_fill = PatternFill("solid", fgColor="DCFCE7")
    scaduto_fill = PatternFill("solid", fgColor="FEE2E2")
    ns_fill = PatternFill("solid", fgColor="FEF9C3")
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 16
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        stato = row[4].value if row[4].value else ""
        if "Aperto" in stato:
            fill = aperto_fill
        elif "Scaduto" in stato:
            fill = scaduto_fill
        else:
            fill = ns_fill
        for cell in row:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row[0].row].height = 30
    ws2 = wb.create_sheet("Riepilogo")
    ws2["A1"] = "Lumen Scout — Monitor Bandi e Finanziamenti"
    ws2["A1"].font = Font(bold=True, size=14, color="0F6E56")
    ws2["A3"] = "Ultimo aggiornamento:"
    ws2["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws2["A3"].font = Font(bold=True)
    ws2["A5"] = "Fonti monitorate:"
    ws2["A5"].font = Font(bold=True)
    ws2["A6"] = "• Invitalia"
    ws2["A7"] = "• Regione Campania - Agricoltura"
    ws2["A8"] = "• GAL Cilento"
    ws2["A9"] = "• UE - Funding & Tenders Portal"
    ws2["A10"] = "• incentivi.gov.it"
    ws2["A12"] = "Totale bandi:"
    ws2["B12"] = n_aperti + n_ns + n_scaduti
    ws2["A13"] = "✅ Aperti:"
    ws2["B13"] = n_aperti
    ws2["A13"].font = Font(color="166534")
    ws2["A14"] = "⏳ Non specificati:"
    ws2["B14"] = n_ns
    ws2["A15"] = "❌ Scaduti:"
    ws2["B15"] = n_scaduti
    ws2["A15"].font = Font(color="991B1B")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20
    wb.save(filename)

# === MAIN ===
print("=" * 50)
print("LUMEN SCOUT — MONITOR BANDI")
print(f"Avvio: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
print("=" * 50)

print("\nScaricando bandi da Invitalia...")
bandi_invitalia = scrapa_invitalia()
print(f"  → {len(bandi_invitalia)} bandi trovati")

print("Scaricando bandi da Regione Campania...")
bandi_campania = scrapa_regione_campania()
print(f"  → {len(bandi_campania)} bandi trovati")

print("Scaricando bandi da GAL Cilento...")
bandi_gal = scrapa_gal_cilento()
print(f"  → {len(bandi_gal)} bandi trovati")

print("Scaricando bandi da incentivi.gov.it...")
bandi_incentivi = scrapa_incentivi_gov()
print(f"  → {len(bandi_incentivi)} bandi trovati")

print("Scaricando bandi da portale UE...")
bandi_ue = scrapa_bandi_ue()
print(f"  → {len(bandi_ue)} bandi trovati")

tutti_i_bandi = bandi_invitalia + bandi_campania + bandi_gal + bandi_incentivi + bandi_ue
print(f"\nTotale bandi raccolti: {len(tutti_i_bandi)}")

df = pd.DataFrame(tutti_i_bandi)
df["Stato"] = df["Scadenza"].apply(calcola_stato)
df["_data_ord"] = df["Scadenza"].apply(parse_data)

# Analisi AI solo sui bandi aperti
print("\nAnalisi AI dei bandi aperti con Claude...")
tipo_beneficiario_list = []
settore_ateco_list = []
settore_ets_list = []
fascia_demografica_list = []
qualita_bando_list = []
motivazione_list = []

aperti_mask = df["Stato"] == "✅ Aperto"
totale_aperti = aperti_mask.sum()

for i, (idx, row) in enumerate(df.iterrows()):
    if row["Stato"] == "✅ Aperto":
        if i % 10 == 0:
            print(f"  Analizzando bando {i+1}/{totale_aperti}...")
        analisi = analizza_bando_con_claude(row["Titolo"], row["Fonte"], row["Scadenza"])
        tipo_beneficiario_list.append(analisi.get("tipo_beneficiario", "N/D"))
        settore_ateco_list.append(analisi.get("settore_ateco", "N/A"))
        settore_ets_list.append(analisi.get("settore_ets", "N/A"))
        fascia_demografica_list.append(analisi.get("fascia_demografica_target", "N/A"))
        qualita_bando_list.append(analisi.get("qualita_bando", "N/D"))
        motivazione_list.append(analisi.get("motivazione", "N/D"))
    else:
        tipo_beneficiario_list.append("")
        settore_ateco_list.append("")
        settore_ets_list.append("")
        fascia_demografica_list.append("")
        qualita_bando_list.append("")
        motivazione_list.append("")

df["Tipo Beneficiario"] = tipo_beneficiario_list
df["Settore ATECO"] = settore_ateco_list
df["Settore ETS"] = settore_ets_list
df["Fascia Demografica"] = fascia_demografica_list
df["Qualità Bando"] = qualita_bando_list
df["Motivazione AI"] = motivazione_list

df_aperti = df[df["Stato"] == "✅ Aperto"].sort_values(
    ["Qualità Bando", "_data_ord"],
    key=lambda x: x.map({"Alta": 0, "Media": 1, "Bassa": 2}) if x.name == "Qualità Bando" else x
)
df_ns = df[df["Stato"] == "Non specificata"]
df_scaduti = df[df["Stato"] == "❌ Scaduto"].sort_values("_data_ord", ascending=False)

df_finale = pd.concat([df_aperti, df_ns, df_scaduti]).drop(columns=["_data_ord"])

filename = "bandi_campania.xlsx"
df_finale.to_excel(filename, index=False)
formatta_excel(filename, len(df_aperti), len(df_ns), len(df_scaduti))

# Notifica email bandi nuovi (pubblicati nelle ultime 48 ore)
oggi = datetime.today()
limite = oggi - timedelta(hours=48)
bandi_aperti_lista = df_aperti.to_dict("records")
bandi_nuovi = []

for b in bandi_aperti_lista:
    data_pub = b.get("Data pubblicazione", "Non specificata")
    if data_pub and data_pub != "Non specificata":
        try:
            dt = datetime.strptime(data_pub, "%d/%m/%Y")
            if dt >= limite:
                bandi_nuovi.append(b)
        except:
            pass

if bandi_nuovi:
    print(f"\n📧 Trovati {len(bandi_nuovi)} bandi pubblicati nelle ultime 48h — invio email...")
    invia_email_bandi(bandi_nuovi, len(df_aperti))
else:
    print("\n📭 Nessun nuovo bando pubblicato nelle ultime 48 ore.")

print(f"\n✅ File Excel salvato con analisi AI!")
print(f"   Aperti: {len(df_aperti)} | Non specificati: {len(df_ns)} | Scaduti: {len(df_scaduti)}")
print(f"\nCompletato: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
print("=" * 50)
